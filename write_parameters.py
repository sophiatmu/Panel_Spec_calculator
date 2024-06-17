# write_parameters.py
import config
import openpyxl


def save_workbook(save_wb):
    workbook = config.workbook
    workbook.save(save_wb)
def write_parameters_to_excel(save_wb):
    for i in range(3):
        config.home_sheet.cell(row=2, column=14 + i).value = config.matrix_max_table[0, i]
        config.home_sheet.cell(row=3, column=14 + i).value = config.matrix_max_table[1, i]
        config.home_sheet.cell(row=4, column=14 + i).value = config.matrix_max_table[2, i]
        config.home_sheet.cell(row=5, column=14 + i).value = config.matrix_max_table[3, i]
        config.home_sheet.cell(row=9, column=14 + i).value = round(config.matrix_max_table[4, i],2)
        # config.home_sheet.cell(row=10, column=14 + i).value = 100*config.matrix_max_table[5, i] #duty
        save_workbook(save_wb)

def Gamma_xlsx(panel_Nits,Grayscale,Max_panel_Nits,save_wb):
    for i in range(3):
        value = panel_Nits[i]/Max_panel_Nits[i]
        # print(float(value))
        config.Gamma_sheet.cell(row=Grayscale+2, column=8+i).value = float(value)

    save_workbook(save_wb)

def Grayscale_xlsx(matrix_max_table,panel_Nits,Grayscale,save_wb):
    for i in range(3):
        ratio = float(panel_Nits[i]/config.Max_Brightness)
        config.Grayscale_sheet.cell(row=Grayscale + 19, column=i * 4 + 2).value = ratio
        config.Grayscale_sheet.cell(row=Grayscale + 19, column=i * 4 + 3).value = matrix_max_table[4, i]
        config.Grayscale_sheet.cell(row=Grayscale + 19, column=i * 4 + 4).value = matrix_max_table[1, i]
        config.Grayscale_sheet.cell(row=Grayscale + 19, column=i * 4 + 5).value = matrix_max_table[3, i]
    save_workbook(save_wb)

def Grayscale_CIE(Color_Space,Grayscale,save_wb):
    for i in range(3):
        config.Grayscale_sheet.cell(row=Grayscale + 1, column=i * 2 + 2).value = Color_Space[1, i] #CIEx
        config.Grayscale_sheet.cell(row=Grayscale + 1, column=i * 2 + 3).value = Color_Space[2, i] #CIEy
    save_workbook(save_wb)