import os
import openpyxl
import numpy as np




def find_optical_files(directory):
    # 找到目錄中所有後綴為 '_B_Optical.xlsx' 的檔案
    R_files = [filename for filename in os.listdir(directory) if filename.endswith('_R_Optical.xlsx')]
    G_files = [filename for filename in os.listdir(directory) if filename.endswith('_G_Optical.xlsx')]
    B_files = [filename for filename in os.listdir(directory) if filename.endswith('_B_Optical.xlsx')]
    return R_files,G_files,B_files


def main():
    directory = '.'  # 可以指定目錄路徑
    R_files,G_files, B_files = find_optical_files(directory)
    # print(optical_files)
    for i in R_files:
        R_wb = openpyxl.load_workbook(i)
        source_R = R_wb['Optical Data']
    for i in G_files:
        G_wb = openpyxl.load_workbook(i)
        source_G = G_wb['Optical Data']
    for i in B_files:
        B_wb = openpyxl.load_workbook(i)
        source_B = B_wb['Optical Data']
    R_font = openpyxl.styles.Font(color='FF0000',name='Arial')
    G_font = openpyxl.styles.Font(color='008000', name='Arial')
    B_font = openpyxl.styles.Font(color='0000ff', name='Arial')

    for col in range(1, source_R.max_column + 1):
        if source_R.cell(row=1, column=col).value == 'Voltage (V)':
            # 複製整行並貼到 summary_ws 中的指定位置
            for row in range(2, source_R.max_row + 1):
                OpticalData_sheet.cell(row=row, column=1).value = source_R.cell(row=row, column=col).value
                OpticalData_sheet.cell(row=row, column=1).font = R_font
        if source_R.cell(row=1, column=col).value == 'Current/chip (uA)':
            for row in range(2, source_R.max_row + 1):
                OpticalData_sheet.cell(row=row, column=2).value = source_R.cell(row=row, column=col).value
                OpticalData_sheet.cell(row=row, column=2).font = R_font
        if source_R.cell(row=1, column=col).value == 'Yield (cd/A)':
            for row in range(2, source_R.max_row + 1):
                OpticalData_sheet.cell(row=row, column=3).value = source_R.cell(row=row, column=col).value
                OpticalData_sheet.cell(row=row, column=3).font = R_font

    for col in range(1, source_G.max_column + 1):
        if source_G.cell(row=1, column=col).value == 'Voltage (V)':
            # 複製整行並貼到 summary_ws 中的指定位置
            for row in range(2, source_G.max_row + 1):
                OpticalData_sheet.cell(row=row, column=6).value = source_G.cell(row=row, column=col).value
                OpticalData_sheet.cell(row=row, column=6).font = G_font
        if source_G.cell(row=1, column=col).value == 'Current/chip (uA)':
            for row in range(2, source_G.max_row + 1):
                OpticalData_sheet.cell(row=row, column=7).value = source_G.cell(row=row, column=col).value
                OpticalData_sheet.cell(row=row, column=7).font = G_font
        if source_G.cell(row=1, column=col).value == 'Yield (cd/A)':
            for row in range(2, source_G.max_row + 1):
                OpticalData_sheet.cell(row=row, column=8).value = source_G.cell(row=row, column=col).value
                OpticalData_sheet.cell(row=row, column=8).font = G_font

    for col in range(1, source_B.max_column + 1):
        if source_B.cell(row=1, column=col).value == 'Voltage (V)':
            # 複製整行並貼到 summary_ws 中的指定位置
            for row in range(2, source_B.max_row + 1):
                OpticalData_sheet.cell(row=row, column=11).value = source_B.cell(row=row, column=col).value
                OpticalData_sheet.cell(row=row, column=11).font = B_font
        if source_B.cell(row=1, column=col).value == 'Current/chip (uA)':
            for row in range(2, source_B.max_row + 1):
                OpticalData_sheet.cell(row=row, column=12).value = source_B.cell(row=row, column=col).value
                OpticalData_sheet.cell(row=row, column=12).font = B_font
        if source_B.cell(row=1, column=col).value == 'Yield (cd/A)':
            for row in range(2, source_B.max_row + 1):
                OpticalData_sheet.cell(row=row, column=13).value = source_B.cell(row=row, column=col).value
                OpticalData_sheet.cell(row=row, column=13).font = B_font


    for i in range(2, 203):
        if R_spectrum.cell(row=1, column=1).value == OpticalData_sheet.cell(row=i, column=1).value:
            voltagerow_R = i #填入的optical位置
        if G_spectrum.cell(row=1, column=1).value == OpticalData_sheet.cell(row=i, column=6).value:
            voltagerow_G = i #填入的optical位置
        if B_spectrum.cell(row=1, column=1).value == OpticalData_sheet.cell(row=i, column=11).value:
            voltagerow_B = i #填入的optical位置

    for i in range(1, R_spectrum.max_column+1):
        temp = np.zeros(3)
        for j in range(2, 203):
            temp[0] = temp[0] + R_spectrum.cell(row=j, column=i).value * CIE_sheet.cell(row=j, column=1).value
            temp[1] = temp[1] + R_spectrum.cell(row=j, column=i).value * CIE_sheet.cell(row=j, column=2).value
            temp[2] = temp[2] + R_spectrum.cell(row=j, column=i).value * CIE_sheet.cell(row=j, column=3).value
        #第一排算完  vol = 1.6
        ciex = temp[0] / (temp[0] + temp[1] + temp[2])
        ciey = temp[1] / (temp[0] + temp[1] + temp[2])
        OpticalData_sheet.cell(row=voltagerow_R+i-1, column=4).value = ciex
        OpticalData_sheet.cell(row=voltagerow_R+i-1, column=5).value = ciey
        OpticalData_sheet.cell(row=voltagerow_R+i-1, column=4).font = R_font
        OpticalData_sheet.cell(row=voltagerow_R+i-1, column=5).font = R_font

    for i in range(1, G_spectrum.max_column+1):
        temp = np.zeros(3)
        for j in range(2, 203):
            temp[0] = temp[0] + G_spectrum.cell(row=j, column=i).value * CIE_sheet.cell(row=j, column=1).value
            temp[1] = temp[1] + G_spectrum.cell(row=j, column=i).value * CIE_sheet.cell(row=j, column=2).value
            temp[2] = temp[2] + G_spectrum.cell(row=j, column=i).value * CIE_sheet.cell(row=j, column=3).value

        ciex = temp[0] / (temp[0] + temp[1] + temp[2])
        ciey = temp[1] / (temp[0] + temp[1] + temp[2])
        OpticalData_sheet.cell(row=voltagerow_G+i-1, column=9).value = ciex
        OpticalData_sheet.cell(row=voltagerow_G+i-1, column=10).value = ciey
        OpticalData_sheet.cell(row=voltagerow_G+i-1, column=9).font = G_font
        OpticalData_sheet.cell(row=voltagerow_G+i-1, column=10).font = G_font

    for i in range(1, B_spectrum.max_column+1):
        temp = np.zeros(3)
        for j in range(2, 203):
            temp[0] = temp[0] + B_spectrum.cell(row=j, column=i).value * CIE_sheet.cell(row=j, column=1).value
            temp[1] = temp[1] + B_spectrum.cell(row=j, column=i).value * CIE_sheet.cell(row=j, column=2).value
            temp[2] = temp[2] + B_spectrum.cell(row=j, column=i).value * CIE_sheet.cell(row=j, column=3).value

        ciex = temp[0] / (temp[0] + temp[1] + temp[2])
        ciey = temp[1] / (temp[0] + temp[1] + temp[2])
        OpticalData_sheet.cell(row=voltagerow_B+i-1, column=14).value = ciex
        OpticalData_sheet.cell(row=voltagerow_B+i-1, column=15).value = ciey
        OpticalData_sheet.cell(row=voltagerow_B+i-1, column=14).font = B_font
        OpticalData_sheet.cell(row=voltagerow_B+i-1, column=15).font = B_font

    # 保存 summary.xlsx
    wb.save(save_wb) # X098_v4_0613.xlsx


if __name__ == "__main__":
    wb = openpyxl.load_workbook('main_excel.xlsm')

    name = input("Enter saving file name whihout '.xlsx' : ")
    save_wb = str(name)+'.xlsx'
    print(save_wb)

    OpticalData_sheet = wb['OpticalData']
    R_spectrum = wb['RSpectrum']
    G_spectrum = wb['GSpectrum']
    B_spectrum = wb['BSpectrum']
    CIE_sheet = wb['CIE1931']

    main()
    input("Press Enter to exit...")