import numpy as np
from openpyxl import load_workbook
import config
import sys
import write_parameters


class MaxCurrentMaxYieldCalculator:
    def __init__(self):
        config.init_parameters()
        self.workbook = config.workbook
        self.OpticalData_sheet = config.OpticalData_sheet
        self.home_sheet = config.home_sheet
        self.Grayscale_sheet = config.Grayscale_sheet
        self.Gamma_sheet = config.Gamma_sheet
        self.R_spectrum = config.R_spectrum
        self.G_spectrum = config.G_spectrum
        self.B_spectrum = config.B_spectrum
        self.Spectrums = config.Spectrums
        self.CFData_sheet = config.CFData_sheet
        self.CIE_sheet = config.CIE_sheet

    def set_parameters(self):
        config.PixelSize = config.home_sheet.cell(row=2, column=2).value
        config.VResolution = config.home_sheet.cell(row=3, column=2).value
        config.HResolution = config.home_sheet.cell(row=4, column=2).value
        config.PanelShape = config.home_sheet.cell(row=5, column=2).value
        config.CPL = config.home_sheet.cell(row=3, column=5).value

        config.Gamma = config.home_sheet.cell(row=8, column=5).value
        config.User_grayscale = config.home_sheet.cell(row=10, column=10).value / 16
        config.Max_Brightness = config.home_sheet.cell(row=1, column=5).value
        config.EM_Brightness = round(config.Max_Brightness * (config.User_grayscale / 16) ** config.Gamma, 2)

        for i in range(3):
            config.Number[i] = config.home_sheet.cell(row=5, column=5 + i).value
            config.Efficiency[i] = config.home_sheet.cell(row=4, column=5 + i).value

        if config.home_sheet.cell(row=2, column=10).value == 0.313 and config.home_sheet.cell(row=2,
                                                                                          column=11).value == 0.329:
            config.ArrayW[0, 0] = 0.9514
            config.ArrayW[1, 0] = 1
            config.ArrayW[2, 0] = 1.0881

        if config.home_sheet.cell(row=7, column=5).value == 2.5:
            config.Beta[0] = 1
            config.Beta[1] = 1
            config.Beta[2] = 1
            if config.home_sheet.cell(row=7, column=6).value == "R Shared":
                config.Beta[0] = 0.5
            elif config.home_sheet.cell(row=7, column=6).value == "G Shared":
                config.Beta[1] = 0.5
            elif config.home_sheet.cell(row=7, column=6).value == "B Shared":
                config.Beta[2] = 0.5
        elif config.home_sheet.cell(row=7, column=5).value == 2:
            config.Beta[0] = 0.5
            config.Beta[1] = 0.5
            config.Beta[2] = 0.5
            if config.home_sheet.cell(row=7, column=6).value == "G&B Shared":
                config.Beta[0] = 1
            elif config.home_sheet.cell(row=7, column=6).value == "R&B Shared":
                config.Beta[1] = 1
            elif config.home_sheet.cell(row=7, column=6).value == "R&G Shared":
                config.Beta[2] = 1
        elif config.home_sheet.cell(row=7, column=5).value == 1.5:
            config.home_sheet.cell(row=7, column=6).value = "N/A"
            config.Beta[0] = 0.5
            config.Beta[1] = 0.5
            config.Beta[2] = 0.5
        else:
            config.home_sheet.cell(row=7, column=6).value = "N/A"

    def load_optical_data(self):
        for i in range(3):
            for row in range(110):  # 2~101
                if config.OpticalData_sheet.cell(row=row + 2, column=i + 1).value:
                    config.R_optical[row, i] = config.OpticalData_sheet.cell(row=row + 2, column=i + 1).value
                if config.OpticalData_sheet.cell(row=row + 2, column=i + 6).value:
                    config.G_optical[row, i] = config.OpticalData_sheet.cell(row=row + 2, column=i + 6).value
                if config.OpticalData_sheet.cell(row=row + 2, column=i + 11).value:
                    config.B_optical[row, i] = config.OpticalData_sheet.cell(row=row + 2, column=i + 11).value

    def calculate_max_current_yield(self):
        for i in range(3):
            max_yield = np.max(config.opticals[i][:, 2])
            max_current = np.max(config.opticals[i][:, 1])

            row_idx, col_idx = np.where(config.opticals[i] == max_yield)
            config.matrix[i, 0] = max_yield
            config.matrix[i, 1] = config.opticals[i][row_idx, 1]  # max current
            config.matrix[i, 2] = config.opticals[i][row_idx, 0]  # max voltage
            config.home_sheet.cell(row=8, column=14 + i, value=round(max_yield, 2))
            config.home_sheet.cell(row=7, column=14 + i, value=round(max_current, 2))

    def set_CFL(self):
        if config.home_sheet.cell(row=9, column=10).value == "Test_RD-Line":  # RD-Line
            row_start = 3
            row_end = 203
            col_start = 19
            col_end = 21
            config.CFL[0] = 0.88
            config.CFL[1] = 0.82
            config.CFL[2] = 0.79
            return row_start,row_end,col_start,col_end
        elif config.home_sheet.cell(row=9, column=10).value == "RD-Line":
            row_start = 3
            row_end = 203
            col_start = 2
            col_end = 4
            config.CFL[0] = 0.88
            config.CFL[1] = 0.82
            config.CFL[2] = 0.79
            return row_start, row_end, col_start, col_end
        elif config.home_sheet.cell(row=9, column=10).value == "C4A":
            row_start = 3
            row_end = 203
            col_start = 7
            col_end = 9
            config.CFL[0] = 0.89
            config.CFL[1] = 0.86
            config.CFL[2] = 0.81
            return row_start, row_end, col_start, col_end
        elif config.home_sheet.cell(row=9, column=10).value == "No":
            row_start = 0
            row_end = 200
            col_start = 20
            col_end = 22
            config.CFL[0] = 1
            config.CFL[1] = 1
            config.CFL[2] = 1
            return row_start, row_end, col_start, col_end
    def get_matrixCIE_ratio(self):

        for i in range(3):
            for col in range(1, config.Spectrums[i].max_column + 1):
                cell_value = config.Spectrums[i].cell(row=1, column=col).value
                next_cell_value = config.Spectrums[i].cell(row=1, column=col + 1).value
                if cell_value == config.matrix[i, 2]:
                    for j in range(201):
                        config.voltage_data[j, i] = config.Spectrums[i].cell(row=j + 2, column=col).value
                        config.temp_matrix[j, i] = config.voltage_data[j, i] * config.color_filter[j, i]

                    break
                elif isinstance(next_cell_value, (int, float)) and cell_value < config.matrix[i, 2] < next_cell_value:
                    for j in range(201):
                        y1 = config.Spectrums[i].cell(row=j + 2, column=col).value  # column1 的值
                        y2 = config.Spectrums[i].cell(row=j + 2, column=col + 1).value  # column2 的值
                        config.voltage_data[j, i] = y1 + (y2 - y1) * (config.matrix[i, 2] - cell_value) / (
                                next_cell_value - cell_value)
                        config.temp_matrix[j, i] = config.voltage_data[j, i] * config.color_filter[j, i]
                    break
            temp1 = 0
            temp2 = 0
            temp3 = 0
            for j in range(201):
                temp1 = round(temp1 + config.temp_matrix[j, i] * config.Color_Matching[j, 0], 4)
                temp2 = round(temp2 + config.temp_matrix[j, i] * config.Color_Matching[j, 1], 4)
                temp3 = round(temp3 + config.temp_matrix[j, i] * config.Color_Matching[j, 2], 4)
            # ----------------------------------------------------------
            # Color_Space = matrixCIE
            # ----------------------------------------------------------
            config.Color_Space[1, i] = temp1 / (temp1 + temp2 + temp3)
            config.Color_Space[2, i] = temp2 / (temp1 + temp2 + temp3)
            config.MatrixRGB[0, i] = temp1 / temp2
            config.MatrixRGB[1, i] = 1
            config.MatrixRGB[2, i] = temp3 / temp2

    def get_middle_data(self,i,panel_Nits):
        for j in range(100):
            cell_value = config.opticals[i][j, 3]
            next_cell_value = config.opticals[i][j + 1, 3]
            next = j
            if (cell_value == panel_Nits[i]):
                cal_yield = config.opticals[i][j, 2]
                cal_current = config.opticals[i][j, 1]
                cal_voltage = config.opticals[i][j, 0]
            elif (cell_value < panel_Nits[i] < next_cell_value):
                x1 = config.opticals[i][j, 2]
                x2 = config.opticals[i][j + 1, 2]
                cal_yield = x1 + (x2 - x1) * (panel_Nits[i] - cell_value) / (next_cell_value - cell_value)
                y1 = config.opticals[i][j, 1]
                y2 = config.opticals[i][j + 1, 1]
                cal_current = y1 + (y2 - y1) * (panel_Nits[i] - cell_value) / (next_cell_value - cell_value)
                z1 = config.opticals[i][j, 0]
                z2 = config.opticals[i][j + 1, 0]
                cal_voltage = z1 + (z2 - z1) * (panel_Nits[i] - cell_value) / (next_cell_value - cell_value)
                break
            elif panel_Nits[i] > np.max(config.opticals[i][:, 3]):
                input("~~~~error：電流計算結果超過 Optical Data 範圍~~~")
                sys.exit()

        return cal_yield,cal_current,cal_voltage,next

    def update_3_data(self,i,cal_yield,cal_current,cal_voltage,panel_Nits,next):

        config.opticals[i][next + 2:, :] = config.opticals[i][next + 1:-1, :]  # 全部下移
        config.opticals[i][next + 1, 0] = cal_voltage
        config.opticals[i][next + 1, 1] = cal_current
        config.opticals[i][next + 1, 2] = cal_yield
        config.opticals[i][next + 1, 3] = panel_Nits[i]

        config.matrix_max_table[0, i] = panel_Nits[i]
        config.matrix_max_table[1, i] = cal_current
        config.matrix_max_table[2, i] = cal_current
        config.matrix_max_table[3, i] = cal_yield
        config.matrix_max_table[4, i] = cal_voltage
        config.matrix_max_table[5, i] = config.dutyRGB[i]


    def overduty(self,panel_Nits,dutyRGB,flag,save_wb):
        dutyRGB = config.dutyRGB

        for i in range(3):
            if flag[i] == 1:
                for j in range(110):
                    config.opticals[i][j, 3] = (config.opticals[i][j, 1] * 10 ** (-6) * config.opticals[i][j, 2] *
                                                config.dutyRGB[i] * config.Beta[i] *
                                                config.Number[i] * config.CPL * config.Efficiency[i] * config.CFL[i]) / (
                                                           config.PixelSize * 10 ** (-6)) ** 2

                # --------------------------------------------
                # panel_Nits 內插 voltage current yield
                # ---------------------------------------------
                cal_yield,cal_current,cal_voltage,next = self.get_middle_data(i,panel_Nits)

                self.update_3_data(i, cal_yield, cal_current, cal_voltage, panel_Nits,next)
                # --------------------------------------------
                # 更新matrix
                # ---------------------------------------------
                config.matrix[i, 0] = cal_yield
                config.matrix[i, 1] = cal_current
                config.matrix[i, 2] = cal_voltage

            #--------------------------------------------------
            # 不用更新 config.matrix
            # --------------------------------------------------
            else:
                config.matrix_max_table[0, i] = panel_Nits[i]
                config.matrix_max_table[1, i] = config.matrix[i, 1] #current
                config.matrix_max_table[2, i] = config.matrix[i, 1] #yield
                config.matrix_max_table[3, i] = config.matrix[i, 0] #voltage
                config.matrix_max_table[4, i] = config.matrix[i, 2]
                config.matrix_max_table[5, i] = config.dutyRGB[i]

        for Iteration in range(4):
            if config.home_sheet.cell(row=1, column=10).value == "Yes":
                # ------------------------------------------------------------
                # 尋找voltage
                # 原始matrixCIE位址
                # ----------------------------------------------------------
                self.get_matrixCIE_ratio()

                config.MatrixRGB = np.linalg.inv(config.MatrixRGB)  # 從這個Matrix 開始有些數值會不一樣
                panel_Nits = np.matmul(config.MatrixRGB, config.ArrayW) * config.Max_Brightness * (
                        config.User_grayscale / 16) ** config.Gamma
            else:
                for i in range(3):
                    panel_Nits[i] = config.home_sheet.cell(row=2, column=14 + i).value

            for i in range(3):
                if flag[i] < 1:
                    # ------------------------------------------------
                    # 只有一開始duty<100 不用迭代 在這邊就算出duty 並直接更新
                    # -----------------------------------------------
                    config.dutyRGB[i] = panel_Nits[i, 0] * (config.PixelSize * 10 ** (-6)) ** 2 / (
                            config.matrix[i, 0] * 10 ** (-6) * config.matrix[i, 1] * config.Beta[i] * config.Number[
                        i] * config.CPL * config.Efficiency[i] * config.CFL[i])


            for i in range(3):
                if flag[i] == 1:
                    # --------------------------------------------
                    # panel_Nits 內插 voltage current yield
                    # ---------------------------------------------
                    cal_yield, cal_current, cal_voltage, next = self.get_middle_data(i, panel_Nits)

                    self.update_3_data(i, cal_yield, cal_current, cal_voltage, panel_Nits, next)
                    # --------------------------------------------
                    # 更新matrix
                    # ---------------------------------------------
                    config.matrix[i, 0] = cal_yield
                    config.matrix[i, 1] = cal_current
                    config.matrix[i, 2] = cal_voltage

                #--------------------------------------------------
                # 不用更新 config.matrix
                # --------------------------------------------------
                else:
                    config.matrix_max_table[0, i] = panel_Nits[i]
                    config.matrix_max_table[1, i] = config.matrix[i, 1] #current
                    config.matrix_max_table[2, i] = config.matrix[i, 1] #yield
                    config.matrix_max_table[3, i] = config.matrix[i, 0] #voltage
                    config.matrix_max_table[4, i] = config.matrix[i, 2]
                    config.matrix_max_table[5, i] = config.dutyRGB[i]
        # ------------------------------------------------------------
        # 算user指定的panel current
        # ----------------------------------------------------------
        panel_current = 0
        for i in range(3):
            panel_current = panel_current + config.dutyRGB[i] * config.Number[i] * config.Beta[i] * \
                            config.matrix[i, 1]

        panel_current = panel_current / 1000000 * config.VResolution * config.HResolution
        if config.PanelShape == "Circle":
            panel_current = panel_current * 3.1415926 / 4
        print(panel_current)
        # ------------------------------------------------------------
        # 寫入panel current 和Em duty
        # ----------------------------------------------------------
        config.home_sheet.cell(row=6, column=14).value = panel_current
        for i in range(3):
            config.home_sheet.cell(row=10, column=14 + i).value = config.dutyRGB[i]
        # ------------------------------------------------------------
        # 開始算最大亮度的panel nits
        # ----------------------------------------------------------
        Max_panel_Nits = np.matmul(config.MatrixRGB, config.ArrayW) * config.Max_Brightness
        write_parameters.Gamma_xlsx(panel_Nits, config.User_grayscale, Max_panel_Nits,save_wb)

        for Grayscale in range(16, 0, -1):
            # if (Grayscale!=config.User_grayscale):
            # -----------------------------------------------------------------
            # 重找每一層voltage Grayscale預設在user時還是迭好迭滿才寫Grayscale表格
            # -----------------------------------------------------------------
            for Iteration in range(4):
                if config.home_sheet.cell(row=1, column=10).value == "Yes":
                    # ------------------------------------------------------------
                    # 尋找voltage
                    # 原始matrixCIE位址
                    # ----------------------------------------------------------
                    self.get_matrixCIE_ratio()

                    config.MatrixRGB = np.linalg.inv(config.MatrixRGB)  # 從這個Matrix 開始有些數值會不一樣
                    panel_Nits = np.matmul(config.MatrixRGB, config.ArrayW) * config.Max_Brightness * (
                                Grayscale / 16) ** config.Gamma
                else:
                    print("non Auto duty")
                for i in range(3):
                    # --------------------------------
                    # 算opticals 中的每個current,yield,voltage對應的各種panelNits 怕超出格子所以用110格
                    # ---------------------------------
                    if Iteration == 0:
                        # opticals只有第一次要重算
                        for j in range(110):
                            config.opticals[i][j, 3] = (config.opticals[i][j, 1] * 10 ** (-6) * config.opticals[i][j, 2] *
                                                        config.dutyRGB[i] * config.Beta[i] *
                                                        config.Number[i] * config.CPL * config.Efficiency[i] * config.CFL[
                                                            i]) / (config.PixelSize * 10 ** (-6)) ** 2

                    # --------------------------------------------
                    # panel_Nits 內插 voltage current yield
                    # ---------------------------------------------
                    cal_yield, cal_current, cal_voltage, next = self.get_middle_data(i, panel_Nits)

                    self.update_3_data(i, cal_yield, cal_current, cal_voltage, panel_Nits, next)
                    # --------------------------------------------
                    # 更新matrix
                    # ---------------------------------------------
                    config.matrix[i, 0] = cal_yield
                    config.matrix[i, 1] = cal_current
                    config.matrix[i, 2] = cal_voltage

            # ------------------------------------
            # 結束迭代 Grayscale那一層最後參數寫入
            # -----------------------------------
            write_parameters.Grayscale_CIE(config.Color_Space, Grayscale,save_wb)
            write_parameters.Gamma_xlsx(panel_Nits, Grayscale, Max_panel_Nits,save_wb)
            write_parameters.Grayscale_xlsx(config.matrix_max_table, panel_Nits, Grayscale,save_wb)
            if Grayscale == 16:
                write_parameters.write_parameters_to_excel(save_wb)
            print(Grayscale)

        workbook = config.workbook
        workbook.save(save_wb)
