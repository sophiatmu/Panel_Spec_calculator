import numpy as np
import openpyxl
from calculate import MaxCurrentMaxYieldCalculator
import config
import write_parameters
import sys

# 模組化減少系統程式
# 打开工作簿和工作表
# wb = load_workbook('D://Muyun//cal_python//X098_v3_0927.xlsx')


def save_workbook():
    workbook = config.workbook
    workbook.save(save_wb)

def main_calculate_flow():
    config.set_workbook(wb)
    config.init_parameters()
    calculator = MaxCurrentMaxYieldCalculator()
    calculator.set_parameters()
    calculator.load_optical_data()
    calculator.calculate_max_current_yield() # 定義config.matrix

    panel_Nits = np.zeros(3)
    if config.home_sheet.cell(row=1, column=10).value == "Yes":
        #-------------------------------------
        # 使用LED Spectrum 所以panel nits要求出
        # -------------------------------------
        vol_column = None
        row_start = 2
        row_end = 202
        col_start = 1
        col_end = 3
        for row in range(row_start, row_end + 1):
            for col in range(col_start, col_end + 1):
                config.Color_Matching[row - row_start][col - col_start] = config.CIE_sheet.cell(row=row, column=col).value

        row_start,row_end,col_start,col_end = calculator.set_CFL()
        if row_start == 0:
            config.color_filter[:][:] = 1
        else:
            for row in range(row_start, row_end + 1):
                for col in range(col_start, col_end + 1):
                    config.color_filter[row - row_start][col - col_start] = config.CFData_sheet.cell(row=row, column=col).value
        # ------------------------------------------------------------
        # 尋找voltage
        # 原始matrixCIE位址
        # ----------------------------------------------------------
        calculator.get_matrixCIE_ratio()

        config.MatrixRGB = np.linalg.inv(config.MatrixRGB)  # 從這個Matrix 開始有些數值會不一樣
        panel_Nits = np.matmul(config.MatrixRGB, config.ArrayW) * config.EM_Brightness

    else:
        #---------------------------------------------
        # not use LED spectrum PanelNits為指定值
        #----------------------------------------------
        for i in range(3):
            panel_Nits[i] = config.home_sheet.cell(row=2, column=14+i).value

    if config.home_sheet.cell(row=2, column=4).value == "Auto Lighting Duty (%)":
        for i in range(3):
            config.dutyRGB[i] = panel_Nits[i, 0] * (config.PixelSize * 10 ** (-6)) ** 2 / (
                        config.matrix[i, 0] * 10 ** (-6) * config.matrix[i, 1] * config.Beta[i] * config.Number[i] * config.CPL * config.Efficiency[i] * config.CFL[i])
            #------------------------------------------------
            # 針對duty>100 不用迭代的值
            #-----------------------------------------------
            if(config.dutyRGB[i]>1):
                config.dutyRGB[i] = 1
                flag[i] = 1
        if 1 in flag:
            config.home_sheet.cell(row=2, column=4).value = "Lighting Duty (%)" #設為非Lighting duty
            calculator.overduty(panel_Nits,config.dutyRGB,flag,save_wb)#整個架構重算 因為要一直迭代
            input("Press Enter to exit...")
            sys.exit()

    else:
        for i in range(3):
            config.dutyRGB[i] = config.home_sheet.cell(row=2, column=5+i).value
    # ------------------------------------------------------------
    # 算user指定的panel current
    # ----------------------------------------------------------
    panel_current = 0
    for i in range(3):
        panel_current = panel_current + config.dutyRGB[i] * config.Number[i] * config.Beta[i] * \
                        config.matrix[i, 1]

    panel_current = panel_current / 1000000 * config.VResolution * config.HResolution
    if config.PanelShape == "Circle":
        panel_current = panel_current * 3.1415926 / 4
    print(panel_current)
    # ------------------------------------------------------------
    # 寫入panel current 和Em duty
    # ----------------------------------------------------------
    config.home_sheet.cell(row=6, column=14).value = panel_current
    for i in range(3):
        config.home_sheet.cell(row=10, column=14 + i).value = config.dutyRGB[i]
    # ------------------------------------------------------------
    # 開始算最大亮度的panel nits
    # ----------------------------------------------------------
    Max_panel_Nits = np.matmul(config.MatrixRGB, config.ArrayW) * config.Max_Brightness
    write_parameters.Gamma_xlsx(panel_Nits, config.User_grayscale, Max_panel_Nits,save_wb)


    for Grayscale in range(16, 0, -1):
        # if (Grayscale!=config.User_grayscale):
        # -----------------------------------------------------------------
        # 重找每一層voltage Grayscale預設在user時還是迭好迭滿才寫Grayscale表格
        # -----------------------------------------------------------------
        for Iteration in range(4):
            if config.home_sheet.cell(row=1, column=10).value == "Yes":
                # ------------------------------------------------------------
                # 尋找voltage
                # 原始matrixCIE位址
                # ----------------------------------------------------------
                calculator.get_matrixCIE_ratio()

                config.MatrixRGB = np.linalg.inv(config.MatrixRGB)  # 從這個Matrix 開始有些數值會不一樣
                panel_Nits = np.matmul(config.MatrixRGB, config.ArrayW) * config.Max_Brightness * (Grayscale / 16) ** config.Gamma
            else:
                print("不使用spctrum")
            for i in range(3):
                # --------------------------------
                # 算opticals 中的每個current,yield,voltage對應的各種panelNits 怕超出格子所以用110格
                # ---------------------------------
                if Iteration == 0:
                    #opticals只有第一次要重算
                    for j in range(110):
                        config.opticals[i][j, 3] = (config.opticals[i][j, 1] * 10 ** (-6) * config.opticals[i][j, 2] * config.dutyRGB[i] * config.Beta[i] *
                                             config.Number[i] * config.CPL * config.Efficiency[i] * config.CFL[i]) / (config.PixelSize * 10 ** (-6)) ** 2

                # --------------------------------------------
                # panel_Nits 內插 voltage current yield
                # ---------------------------------------------
                cal_yield, cal_current, cal_voltage, next = calculator.get_middle_data(i, panel_Nits)

                calculator.update_3_data(i, cal_yield, cal_current, cal_voltage, panel_Nits, next)
                # --------------------------------------------
                # 更新matrix
                # ---------------------------------------------
                config.matrix[i, 0] = cal_yield
                config.matrix[i, 1] = cal_current
                config.matrix[i, 2] = cal_voltage


        # ------------------------------------
        # 結束迭代 Grayscale那一層最後參數寫入
        # -----------------------------------
        write_parameters.Grayscale_CIE(config.Color_Space, Grayscale,save_wb)
        write_parameters.Gamma_xlsx(panel_Nits, Grayscale,Max_panel_Nits,save_wb)
        write_parameters.Grayscale_xlsx(config.matrix_max_table,panel_Nits, Grayscale,save_wb)
        if Grayscale == 16:
            write_parameters.write_parameters_to_excel(save_wb)
        print(Grayscale)


    # 保存工作簿
    # wb.save('D://Muyun//cal_python//X098_v5_0606.xlsx')
    save_workbook()

if __name__=='__main__':
    wb = openpyxl.load_workbook('main_excel.xlsm')
    name = input("Enter saving file name whihout '.xlsx' : ")
    save_wb = str(name)+'.xlsx'
    print(save_wb)
    flag = np.zeros(3)  # 針對rgb duty>100做處理 平常為0
    main_calculate_flow()
    input("Press Enter to exit...")